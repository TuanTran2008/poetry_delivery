import shutil
import glob
import os
import os.path as osp
from datetime import datetime

from svn import library as svn_lib

try:
    import openpyxl
except:
    svn_lib.addsitedir("python_module/openpyxl-2.1.2-py2.7")
    import openpyxl
from openpyxl.styles import Style, Alignment

ffmpeg_bin = svn_lib.require('bin/ffmpeg2018.N89936/ffmpeg.exe')

from shot.workflow.util import pathfunc
import clarity

CLIENT_KEYWORD = dict(
    FOLDER_CLIENT="mc_pdp",
    SCENES="scenes",
    ANIMATION="Animation",
    SEASON="S_03",
    EDIT="_editing",
    LOCAL="WIP",
    AV="Av",
    ANIM="Anim",
    PROJECT="PDP",
    PROJECT_CLIENT="PDT",
    ENV="OUTGOING_DIR",
    SPARX="SPR",
    DATE=datetime.now().strftime("%d%m%y")
)
WORKSPACE_CLIENT = os.environ["MAYACONTENT"]

FOLDER_TEMP = osp.join(os.environ["TEMP"], "r_sync_PDP_{date}".format(date=datetime.now().strftime("%Y%m%d_%H%M%S")))

DELI_SHOT_TXT = r"${ENV\deliveredShot.txt"
DELI_MOV_TXT = r"${ENV}\deliveredMov.txt"
DELI_SHOT_XLS = r"deliveredMov{date}.xlsx"



FIRST_MOV_PATH = r"${ENV}\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{EDIT}\{step_client}"
SECOND_MOV_PATH = r"${ENV}\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{EDIT}"
THIRD_MOV_PATH = r"${ENV}\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{seq}_{shot}\{LOCAL}\{AV}"

MAYA_CLIENT_PATH = "${ENV}\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{seq}_{shot}\{LOCAL}\{step_client}"

FIRST_FILE_MOV = "{PROJECT}_{eps}_{seq}_{shot}_{ANIM}_{step_client}_v{version}.mov"
SECOND_FILE_MOV = "{PROJECT}_{eps}_{seq}_{shot}.mov"
THIRD_FILE_MOV = "{PROJECT_CLIENT}_{eps}_{shot}_{ANIM}_{step_client}_v{version}_{DATE}.mov"

MAYA_CLIENT_FILE = "PDP_{eps}_{seq}_{shot}_{ANIM}_{step_client}_v{version}_{SPARX}.ma"

MOV_EXT = ".mov"


def map_shotcode(shot_code, return_client=True):
    eps, seq, shot = shot_code.split(".")

    if return_client:
        if eps.lower() in ["306a", "306b"]:
            eps = eps.upper()
            seq = str(int(seq)).zfill(3)
            shot = shot.upper()

    return eps, seq, shot


def _get_media_file(shot_code, step):
    msg = None
    eps, seq, shot = shot_code.split(".")
    if eps in ["306a", "306b"]:
        eps_sparx = eps.lower()
        seq_sparx = str(int(seq)).zfill(2)
    else:
        eps_sparx = eps
        seq_sparx = seq

    last_version_dir, _ = pathfunc.get_last_version_image(eps_sparx, seq_sparx, shot, step, pathfunc.IS_REVIEW)
    shot_mov_patt = pathfunc.shot_mov_patt.format(eps=eps_sparx, seq=seq_sparx, shot=shot).replace("#", "*")
    shot_jpg_patt = pathfunc.shot_jpg_patt.format(eps=eps_sparx, seq=seq_sparx, shot=shot).replace("#", "%04d")
    audio_file = pathfunc.get_aif_last(eps, seq, shot) or pathfunc.get_wav_last(eps, seq, shot)

    if not osp.isfile(audio_file):
        msg = "Not exists audio"

    if not osp.isdir(last_version_dir):
        msg = "Not exists Movie file"
        raise IOError(msg)

    mov_file = glob.glob(osp.join(last_version_dir, shot_mov_patt))[0]
    return msg, mov_file, osp.join(last_version_dir, shot_jpg_patt), audio_file


def get_mov_client(shot_code, step):
    ref_mov = """
    	First place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\_editing\Lay\PDP_305A_002_040_Anim_Lay_v003.mov  
    	(here the version number is incremental in order to keep track of all the history)

    	Second place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\_editing\PDP_305A_002_040.mov
    	(here the file name is always the same in order to be taken automatically by our editors) 

    	Third place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\002_040\WIP\Av\PDT_305A_040_Anim_Lay_v003_{date}.mov

    	"""
    msg, mov_file, jpg_file, audio_file = _get_media_file(shot_code, step)
    eps, seq, shot = map_shotcode(shot_code)
    clone_client_keyword = CLIENT_KEYWORD.copy()
    dirname, base_name = osp.split(mov_file)
    last_version = osp.basename(osp.dirname(dirname))
    last_version = str(int(last_version)).zfill(3)

    if step == "layout":
        step_client = "Lay"
    else:
        step_client = "Sec"
    clone_client_keyword.update(dict(step_client=step_client, version=last_version,
                                     eps=eps, seq=seq, shot=shot, step=step.title()))

    first_place_template = clarity.Template("first_mov", osp.join(FIRST_MOV_PATH, FIRST_FILE_MOV))
    second_place_template = clarity.Template("second_mov", osp.join(SECOND_MOV_PATH, SECOND_FILE_MOV))
    third_place_template = clarity.Template("third_mov", osp.join(THIRD_MOV_PATH, THIRD_FILE_MOV))

    first_place = first_place_template.format(clone_client_keyword, expand_var=True)
    second_place = second_place_template.format(clone_client_keyword, expand_var=True)
    third_place = third_place_template.format(clone_client_keyword, expand_var=True)

    # Remove all scene not lastest version
    # clone_client_keyword.update(version="*")
    # first_previous_version = first_place_template.format(clone_client_keyword, expand_var=True)
    # second_previous_version = second_place_template.format(clone_client_keyword, expand_var=True)
    # third_previous_version = third_place_template.format(clone_client_keyword, expand_var=True)
    #
    # first_previous_lst = glob.glob(first_previous_version)
    # second_previous_lst = glob.glob(second_previous_version)
    # third_previous_lst = glob.glob(third_previous_version)
    #
    # lst_previous = first_previous_lst + second_previous_lst + third_previous_lst
    #
    # if lst_previous:
    #     print "Trying remove previous movie of {}".format(".".join([eps, seq, shot]))
    #     for file_mov in lst_previous:
    #         print file_mov
    #         os.remove(file_mov)

    return mov_file, first_place, second_place, third_place


def get_maya_temp(maya_file):
    file_name = osp.basename(maya_file)
    if not osp.isdir(FOLDER_TEMP):
        os.makedirs(FOLDER_TEMP)
    return osp.join(FOLDER_TEMP, file_name)


def get_maya_client(shot_code, step, scene_file):
    doc = """
    	"P:\mc_pdp\scenes\Animation\S_03\300\000_000\WIP\Lay\PDP_300_000_000_Anim_Lay_v008_SPR.ma"
    	"P:\mc_pdp\scenes\Animation\S_03\300\000_000\WIP\Sec\PDP_300_000_000_Anim_Sec_v001_SPR.ma"
    	"""

    # Get lastest version of movie
    mov_file = get_mov_client(shot_code, step)
    mov_file = osp.dirname(mov_file[0])
    last_version_mov = osp.basename(osp.dirname(mov_file))

    eps, seq, shot = map_shotcode(shot_code)

    clone_client_keyword = CLIENT_KEYWORD.copy()
    dirname, base_name = osp.split(scene_file)
    last_version = base_name.split(".")[4]

    if int(last_version_mov) != int(last_version):
        raise IOError("Cannot delivery because have different version between maya and movie, please check again")

    last_version = str(int(last_version)).zfill(3)

    if step == "layout":
        step_client = "Lay"
    else:
        step_client = "Sec"
    clone_client_keyword.update(dict(step_client=step_client, version=last_version,
                                     eps=eps, seq=seq, shot=shot, step=step.title()))
    first_place_template = clarity.Template("maya_client", osp.join(MAYA_CLIENT_PATH, MAYA_CLIENT_FILE))

    first_place = first_place_template.format(clone_client_keyword, expand_var=True)

    # Remove all scene not lastest version
    # clone_client_keyword.update(version="*")
    # previous_version = first_place_template.format(clone_client_keyword, expand_var=True)
    # lst_previous = glob.glob(previous_version)
    # if lst_previous:
    #     print "Trying remove previous maya of {}".format(shot_code)
    #
    #     for file_maya in lst_previous:
    #         print file_maya
    #         os.remove(file_maya)

    return first_place


def last_version_in_mov_result(file_txt):
    dct = {}
    dct_maya = {}
    if not file_txt:
        return False

    def toBool(x):
        if type(x) is str:
            return x in ["True"]
        else:
            return bool(x)

    with open(file_txt, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                shotcode, version, is_copy_maya = line.split(" ")
            except ValueError:
                shotcode, version = line.split(" ")
                is_copy_maya = "True"

            if dct.get(shotcode, True):
                dct.update({shotcode: version})
                dct_maya.update({shotcode: toBool(is_copy_maya)})
            else:
                if int(version) > int(dct.get(shotcode)):
                    dct.update({shotcode: version})
                    dct_maya.update({shotcode: toBool(is_copy_maya)})

    os.remove(file_txt)

    # create xlss file.
    file_xls = file_txt.replace(".txt", ".xlsx")
    if osp.isfile(file_xls):
        try:
            os.remove(file_xls)
        except WindowsError as e:
            dir_name, base_name = osp.split(file_xls)
            file_xls_new = DELI_SHOT_XLS.format(date=datetime.now().strftime("%Y%m%d_%H%M%S"))
            file_xls = osp.join(dir_name, file_xls_new)
    wb = openpyxl.Workbook()
    date_deli = datetime.now().strftime("%m/%d/%Y")
    if file_xls:
        titlenames = ("Delivery Tracking",)
        headnames=("","","Playblast","","Maya file")
        fieldnames = ('Shot Code', 'Animation', 'Version', 'Sent date', 'Version', 'Sent date')
        sheet = wb.active

        sheet.merge_cells("A1:F1")
        sheet.merge_cells("C2:D2")
        sheet.merge_cells("E2:F2")
        sheet.append(titlenames)
        sheet.append(headnames)
        sheet.append(fieldnames)

        # append all rows
        for row in sorted(dct.keys()):
            if dct_maya[row]:
                tmp = (row, "For review", dct[row], date_deli, dct[row], date_deli)
            else:
                tmp = (row, "For review", dct[row], date_deli, "", "")
            sheet.append(tmp)

        max_row = sheet.max_row
        max_column = sheet.max_column
        for i in range(1, max_row+1):
            for j in range(1, max_column+1):
                cell_obj = sheet.cell(row=i, column=j)
                cell_obj.style = cell_obj.style.copy(alignment=Alignment(horizontal='center'))
        # save file
        wb.save(file_xls)

    with open(file_txt, "w") as f:
        for shotcode in sorted(dct.keys()):
            f.write("{} {} {}\n".format(shotcode, dct[shotcode], dct_maya[shotcode]))


def last_version_in_maya_result(file_txt):
    dct = {}
    if not file_txt:
        return False
    with open(file_txt, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            eps, seq, shot, _, step, version = line.split("_")[1:7]
            shotcode = "_".join([eps, seq, shot, step])
            if dct.get(shotcode, True):
                dct.update({shotcode: version})
            else:
                if int(version) > int(dct.get(shotcode)):
                    dct.update({shotcode: version})

    os.remove(file_txt)
    template = "PDP_{eps}_{seq}_{shot}_Anim_{step}_{version}_SPR.ma"
    with open(file_txt, "w") as f:
        for shotcode in dct.keys():
            eps, seq, shot, step = shotcode.split("_")
            version = dct[shotcode]
            f.write("{}\n".format(template.format(eps=eps, seq=seq, shot=shot,
                                                  step=step, version=version)))
    return True
