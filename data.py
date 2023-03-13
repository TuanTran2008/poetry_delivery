import shutil
import glob
import os
import os.path as osp
from datetime import datetime

from svn import library as svn_lib

try:
    import openpyxl
except:
    svn_lib.addsitedir("python_module/openpyxl-2.1.2-py2.7")
    import openpyxl
from openpyxl.styles import Style, Alignment

ffmpeg_bin = svn_lib.require('bin/ffmpeg2018.N89936/ffmpeg.exe')

from shot.workflow.util import pathfunc
from shot.workflow.util import fn_shotversion
import clarity

# PATH = os.environ["MAYACONTENT"]
OUT_GOING = r"OUTGOING_DIR"
os.environ["OUTGOING_DIR"] = osp.join(os.environ["OUTGOING_DIR"],
                                      datetime.now().strftime("%Y-%m-%d"))
CLIENT_KEYWORD = dict(
    FOLDER_CLIENT="mc_pdp",
    SCENES="scenes",
    ANIMATION="Animation",
    SEASON="S_03",
    EDIT="_editing",
    LOCAL="WIP",
    AV="Av",
    ANIM="Anim",
    PROJECT="PDP",
    PROJECT_CLIENT="PDT",
    ENV="OUTGOING_DIR",
    SPARX="SPR",
    DATE=datetime.now().strftime("%d%m%y"),
    DATE2=datetime.now().strftime("%Y-%m-%d")
)

FOLDER_TEMP = osp.join(os.environ["TEMP"], "delivery_PDP_{date}".format(date=datetime.now().strftime("%Y%m%d_%H%M%S")))

DELI_SHOT_TXT = r"${ENV}\{step}\deliveredShot.txt"
DELI_MOV_TXT = r"${ENV}\{step}\deliveredMov.txt"
DELI_SHOT_XLS = r"deliveredMov{date}.xlsx"
DELI_SHOT_XLS_BK = r"${ENV}\{step}\backup\deliveredMov_{date}.xlsx"

FIRST_MOV_PATH = r"${ENV}\{step}\{DATE2}_{eps}-{department}Delivery\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{EDIT}\{step_client}"
SECOND_MOV_PATH = r"${ENV}\{step}\{DATE2}_{eps}-{department}Delivery\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{EDIT}"
THIRD_MOV_PATH = r"${ENV}\{step}\{DATE2}_{eps}-{department}Delivery\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{seq}_{shot}\{LOCAL}\{AV}"

MAYA_CLIENT_PATH = "${ENV}\{step}\{DATE2}_{eps}-{department}Delivery\{FOLDER_CLIENT}\{SCENES}\{ANIMATION}\{SEASON}\{eps}\{seq}_{shot}\{LOCAL}\{step_client}"

FIRST_FILE_MOV = "{PROJECT}_{eps}_{seq}_{shot}_{ANIM}_{step_client}_v{version}.mov"
SECOND_FILE_MOV = "{PROJECT}_{eps}_{seq}_{shot}.mov"
THIRD_FILE_MOV = "{PROJECT_CLIENT}_{eps}_{shot}_{ANIM}_{step_client}_v{version}_{DATE}.mov"

MAYA_CLIENT_FILE = "PDP_{eps}_{seq}_{shot}_{ANIM}_{step_client}_v{version}_{SPARX}.ma"

MOV_EXT = ".mov"


def _is_lost_animtaion(new_scene):
    with open(new_scene) as f:
        contents = f.read()
    result = "file -r -ns" in contents and "createNode reference" not in contents
    return result


def get_dct(eps, seq, shot, step, version):
    clone_client_keyword = CLIENT_KEYWORD.copy()

    if eps.startswith("4"):
        clone_client_keyword.update(dict(SEASON="S_04"))

    if step == "layout":
        step_client = "Lay"
        department = "Layout"
    else:
        step_client = "Sec"
        department = "Animation"
    clone_client_keyword.update(dict(step_client=step_client, version=version,
                                     eps=eps, seq=seq, shot=shot, step=step.title(), department=department))
    return clone_client_keyword


def _get_media_file(shot_code, step):
    msg = None
    eps, seq, shot = shot_code.split(".")
    if eps in ["306a", "306b"]:
        eps_sparx = eps.lower()
        seq_sparx = str(int(seq)).zfill(2)
    else:
        eps_sparx = eps
        seq_sparx = seq

    last_version_dir, _ = pathfunc.get_last_version_image(eps_sparx, seq_sparx, shot, step, pathfunc.IS_REVIEW)
    shot_mov_patt = pathfunc.shot_mov_patt.format(eps=eps_sparx, seq=seq_sparx, shot=shot).replace("#", "*")
    shot_jpg_patt = pathfunc.shot_jpg_patt.format(eps=eps_sparx, seq=seq_sparx, shot=shot).replace("#", "%04d")
    audio_file = pathfunc.get_aif_last(eps, seq, shot) or pathfunc.get_wav_last(eps, seq, shot)

    if not osp.isfile(audio_file):
        msg = "Not exists audio"

    if not osp.isdir(last_version_dir):
        msg = "Not exists Movie file"
        raise IOError(msg)

    mov_file = glob.glob(osp.join(last_version_dir, shot_mov_patt))[0]
    return msg, mov_file, osp.join(last_version_dir, shot_jpg_patt), audio_file


def _get_mov_client(shot_code, step):
    ref_mov = """
    	First place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\_editing\Lay\PDP_305A_002_040_Anim_Lay_v003.mov  
    	(here the version number is incremental in order to keep track of all the history)

    	Second place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\_editing\PDP_305A_002_040.mov
    	(here the file name is always the same in order to be taken automatically by our editors) 

    	Third place is:
    	P:\mc_pdp\scenes\Animation\S_03\305A\002_040\WIP\Av\PDT_305A_040_Anim_Lay_v003_{date}.mov

    	"""
    msg, mov_file, jpg_file, audio_file = _get_media_file(shot_code, step)
    eps, seq, shot = map_shotcode(shot_code)

    dirname, base_name = osp.split(mov_file)
    last_version = osp.basename(osp.dirname(dirname))
    last_version = str(int(last_version)).zfill(3)

    clone_client_keyword = get_dct(eps, seq, shot, step, last_version)

    first_place_template = clarity.Template("first_mov", osp.join(FIRST_MOV_PATH, FIRST_FILE_MOV))
    second_place_template = clarity.Template("second_mov", osp.join(SECOND_MOV_PATH, SECOND_FILE_MOV))
    third_place_template = clarity.Template("third_mov", osp.join(THIRD_MOV_PATH, THIRD_FILE_MOV))

    first_place = first_place_template.format(clone_client_keyword, expand_var=True)
    second_place = second_place_template.format(clone_client_keyword, expand_var=True)
    third_place = third_place_template.format(clone_client_keyword, expand_var=True)

    # Remove all scene not lastest version
    clone_client_keyword.update(version="*")
    first_previous_version = first_place_template.format(clone_client_keyword, expand_var=True)
    second_previous_version = second_place_template.format(clone_client_keyword, expand_var=True)
    third_previous_version = third_place_template.format(clone_client_keyword, expand_var=True)

    first_previous_lst = glob.glob(first_previous_version)
    second_previous_lst = glob.glob(second_previous_version)
    third_previous_lst = glob.glob(third_previous_version)

    lst_previous = first_previous_lst + second_previous_lst + third_previous_lst

    if lst_previous:
        print "Trying remove previous movie of {}".format(".".join([eps, seq, shot]))
        for file_mov in lst_previous:
            print file_mov
            os.remove(file_mov)

    return mov_file, first_place, second_place, third_place


def _get_maya_temp(maya_file):
    file_name = osp.basename(maya_file)
    if not osp.isdir(FOLDER_TEMP):
        os.makedirs(FOLDER_TEMP)
    return osp.join(FOLDER_TEMP, file_name)


def _get_maya_client(shot_code, step, scene_file):
    doc = """
    	"P:\mc_pdp\scenes\Animation\S_03\300\000_000\WIP\Lay\PDP_300_000_000_Anim_Lay_v008_SPR.ma"
    	"P:\mc_pdp\scenes\Animation\S_03\300\000_000\WIP\Sec\PDP_300_000_000_Anim_Sec_v001_SPR.ma"
    	"""

    # Get lastest version of movie
    mov_file = _get_mov_client(shot_code, step)
    mov_file = osp.dirname(mov_file[0])
    last_version_mov = osp.basename(osp.dirname(mov_file))

    eps, seq, shot = map_shotcode(shot_code)

    dirname, base_name = osp.split(scene_file)
    last_version = base_name.split(".")[4]

    if int(last_version_mov) != int(last_version):
        raise IOError("Cannot delivery because have different version between maya and movie, please check again")

    last_version = str(int(last_version)).zfill(3)

    clone_client_keyword = get_dct(eps, seq, shot, step, last_version)

    first_place_template = clarity.Template("maya_client", osp.join(MAYA_CLIENT_PATH, MAYA_CLIENT_FILE))

    first_place = first_place_template.format(clone_client_keyword, expand_var=True)

    # Remove all scene not lastest version
    clone_client_keyword.update(version="*")
    previous_version = first_place_template.format(clone_client_keyword, expand_var=True)
    lst_previous = glob.glob(previous_version)
    if lst_previous:
        print "Trying remove previous maya of {}".format(shot_code)

        for file_maya in lst_previous:
            print file_maya
            os.remove(file_maya)

    return first_place


def _copy_file(src, dst):
    if not osp.isdir(osp.dirname(dst)):
        os.makedirs(osp.dirname(dst))
    shutil.copyfile(src, dst)


def map_shotcode(shot_code, return_client=True):
    eps, seq, shot = shot_code.split(".")

    if return_client:
        if eps.lower() in ["306a", "306b"]:
            eps = eps.upper()
            seq = str(int(seq)).zfill(3)
            shot = shot.upper()

    return eps, seq, shot


def _copy_movie(shot_code, step):
    # temp_folder = osp.join(os.environ["TEMP"], "mov_convert")
    # convert_file = osp.basename(mov_file)
    # output_mov = osp.join(temp_folder, convert_file)
    # cmdString = _func_make_movie(jpg_file, 1001, 23.976, output_mov, audio_version=audio_file)
    # cmdstring = ' '.join(cmdString)
    # os.system(cmdstring)

    mov_file, first_file, second_file, third_file = _get_mov_client(shot_code, step)
    # dst_file = _get_maya_client(eps, seq, shot,step, mov_file)
    # _copy_file(scene_file, dst_file)
    # if msg:
    #     lst_error.append([shot_code, msg])
    _copy_file(mov_file, first_file)
    _copy_file(mov_file, second_file)
    _copy_file(mov_file, third_file)

    return mov_file, first_file, second_file, third_file


def _copy_maya(shot_code, step):
    import preprocess
    # eps,seq,shot=shot_code.split(".")
    # lst_scene = glob.glob(osp.join(os.environ["REV_SHOT_CENTRAL"], eps, seq, shot, step, "*.ma"))
    # sort_lst = sorted(lst_scene, reverse=True)
    msg = None
    _, scene_file = pathfunc.get_latest_scene_task(pathfunc.IS_REVIEW, shot_code, b_steps=[step],
                                                   exact_step=True)
    if not osp.isfile(scene_file):
        raise IOError("Please check shot node again, Cannot find it in server's folder")
    if _is_lost_animtaion(scene_file):
        raise IOError("This scene got lost animation, Cannot delivery. Please check again")

    log_obj = fn_shotversion._init_from_scene(scene_file)
    if log_obj.qc_exist():
        if "qcBakeTail" in log_obj.qc_info.detail.keys():
            raise IOError("Please bake key into Tail Ctrl before delivery")
    else:
        raise IOError("Please check QC")

    dst_file = _get_maya_client(shot_code, step, scene_file)
    temp_file = _get_maya_temp(scene_file)
    _copy_file(scene_file, temp_file)
    try:
        preprocess.pre_process(temp_file, dst_file)
    except IOError as e:
        msg = "Cannot pre_process " + str(e)

    return msg, dst_file,log_obj


def create_delivered_shot(lst_scene_deliveredShot, mode, is_copy_maya=True):
    lst_anim = []
    lst_layout = []
    for maya_scene in lst_scene_deliveredShot:
        step = maya_scene.split("_")[5]
        if step == "Lay":
            lst_layout.append(maya_scene)
        else:
            lst_anim.append(maya_scene)

    if lst_anim:
        return create_delivered_txt_file(lst_anim, "Anim", mode, is_copy_maya)

    if lst_layout:
        return create_delivered_txt_file(lst_layout, "Layout", mode)


def create_delivered_txt_file(lst, step, mode="maya", have_maya=True):
    if mode == "maya":
        deliveredshot_file_template = clarity.Template("delivered_shot_txt", DELI_SHOT_TXT)
    else:
        deliveredshot_file_template = clarity.Template("delivered_shot_txt", DELI_MOV_TXT)

    CLIENT_KEYWORD.update(step=step)

    deliveredshot_file = deliveredshot_file_template.format(CLIENT_KEYWORD, expand_var=True)

    with open(deliveredshot_file, "a") as f:
        for ele in lst:
            if mode == "maya":
                f.write(ele + "\n")
            else:
                # PDP_305A_002_040_Anim_Lay_v003.mov
                base_name, ext = ele.split(".")
                eps, seq, shot = base_name.split("_")[1:4]
                version = base_name.split("_")[-1]
                f.write("{} {} {}\n".format("_".join([eps, seq, shot]), version, have_maya))

    return deliveredshot_file


def last_version_in_mov_result(file_txt):
    dct = {}
    dct_maya = {}
    if not file_txt:
        return False

    def toBool(x):
        if type(x) is str:
            return x in ["True"]
        else:
            return bool(x)

    with open(file_txt, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                shotcode, version, is_copy_maya = line.split(" ")
            except ValueError:
                shotcode, version = line.split(" ")
                is_copy_maya = "True"
            eps, seq, shot = shotcode.split("_")
            shotcode = "_".join([eps, str(int(seq)).zfill(2), shot])
            if dct.get(shotcode, True):
                dct.update({shotcode: version})
                dct_maya.update({shotcode: toBool(is_copy_maya)})
            else:
                if int(version) > int(dct.get(shotcode)):
                    dct.update({shotcode: version})
                    dct_maya.update({shotcode: toBool(is_copy_maya)})

    os.remove(file_txt)

    # create xlss file.
    file_xls = file_txt.replace(".txt", ".xlsx")
    if osp.isfile(file_xls):
        try:
            os.remove(file_xls)
        except WindowsError as e:
            dir_name, base_name = osp.split(file_xls)
            file_xls_new = DELI_SHOT_XLS.format(date=datetime.now().strftime("%Y%m%d_%H%M%S"))
            file_xls = osp.join(dir_name, file_xls_new)
    if file_xls:
        create_xls_file(file_xls, dct)

    with open(file_txt, "w") as f:
        for shotcode in sorted(dct.keys()):
            eps, seq, shot = shotcode.split("_")
            shotcode_new = "_".join([eps, str(int(seq)).zfill(3), shot])
            f.write("{} {} {}\n".format(shotcode_new, dct[shotcode], dct_maya[shotcode]))


def create_xls_file(file_xls,dct):
    wb = openpyxl.Workbook()
    date_deli = datetime.now().strftime("%m/%d/%Y")
    titlenames = ("Delivery Tracking",)
    headnames = ("",)
    fieldnames = ('Shot Code', 'Animation', 'Sent date', 'Shot Version')
    sheet = wb.active

    sheet.merge_cells("A1:D1")
    sheet.merge_cells("A2:D2")

    sheet.append(titlenames)
    sheet.append(headnames)
    sheet.append(fieldnames)

    # append all rows
    for row in sorted(dct.keys()):
        # if dct_maya[row]:
        #     tmp = (row, "For review", dct[row], date_deli, dct[row], date_deli)
        # else:
        tmp = (row, "For review", date_deli, dct[row])
        sheet.append(tmp)

    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            cell_obj = sheet.cell(row=i, column=j)
            cell_obj.style = cell_obj.style.copy(alignment=Alignment(horizontal='center'))
    # save file
    wb.save(file_xls)


def last_version_in_maya_result(file_txt):
    dct = {}
    if not file_txt:
        return False
    with open(file_txt, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            eps, seq, shot, _, step, version = line.split("_")[1:7]
            shotcode = "_".join([eps, seq, shot, step])
            if dct.get(shotcode, True):
                dct.update({shotcode: version})
            else:
                if int(version) > int(dct.get(shotcode)):
                    dct.update({shotcode: version})

    os.remove(file_txt)
    template = "PDP_{eps}_{seq}_{shot}_Anim_{step}_{version}_SPR.ma"
    with open(file_txt, "w") as f:
        for shotcode in dct.keys():
            eps, seq, shot, step = shotcode.split("_")
            version = dct[shotcode]
            f.write("{}\n".format(template.format(eps=eps, seq=seq, shot=shot,
                                                  step=step, version=version)))
    return True


def run():
    print "\nDelivery  Tool\n"
    help_txt = "Please drag and drop txt file here:\n"
    file_txt = raw_input(help_txt)
    file_txt = file_txt.strip().replace('"', "")

    copy_maya = raw_input("Do you want to delivery maya scene(y/n)(default is yes): ")
    if copy_maya != "n":
        is_copy_maya = True
    lst_error = []
    lst_success = []
    lst_scene_deliveredShot = []
    lst_mov_deliveredShot = []
    lst_log_qc = []
    bk_file = osp.join(os.environ["LOG_EVENT"], "backup",
                       "delivery_{date}".format(date=datetime.now().strftime("%Y%m%d_%H%M%S")), osp.basename(file_txt))
    if not osp.isdir(osp.dirname(bk_file)):
        os.makedirs(osp.dirname(bk_file))
    shutil.copyfile(file_txt, bk_file)
    with open(file_txt.strip()) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                shot_code, step = line.rsplit(".", 1)
                print "Trying Publish from {} file".format(file_txt)
                print shot_code
            except ValueError as e:
                lst_error.append([line, "Cannot parse this line {}".format(line)])
                continue
            try:
                if is_copy_maya:
                    msg, scene_maya,log_objects = _copy_maya(shot_code, step)
                    if msg:
                        raise IOError(msg)
                    lst_scene_deliveredShot.append(osp.basename(scene_maya))

                    if log_objects.qc_exist().qc_code in ["Error","Bug"]:
                        lst_log_qc.append([shot_code, step, log_objects])

                mov_file, first_file, second_file, third_file = _copy_movie(shot_code, step)
                lst_mov_deliveredShot.append(osp.basename(first_file))
                lst_success.append([shot_code, osp.basename(mov_file), first_file, second_file, third_file])

            except Exception as e:
                lst_error.append([shot_code, str(e)])

    LOG_FILE = osp.join(os.environ["LOG_EVENT"], "delivery", "delivery_{date}.txt")
    log_file = LOG_FILE.format(date=datetime.now().strftime("%Y%m%d_%H%M%S"))
    if not osp.isdir(osp.dirname(log_file)):
        os.makedirs(osp.dirname(log_file))
    with open(log_file, "w") as f:
        for i, e in lst_error:
            f.write("Error shot {}\n".format(i))
            f.write(e + "\n")
            f.write("End Error shot {}\n".format(i))
            f.write("\n")
        for shot_code, i, first, s, t in lst_success:
            f.write("Success shot {}\n".format(shot_code))
            f.write(i + "\n")
            f.write(first + "\n")
            f.write(s + "\n")
            f.write(t + "\n")
            f.write("End Success shot {}\n".format(shot_code))
            f.write("\n")

    with open(log_file, "a") as f:
        if lst_log_qc:
            f.write("Details: all shots didn't pass QC\n\n")
        else:
            f.write("All shot passed QC\n")
        for shotcode,step,log_obj in lst_log_qc:
            qc_obj = log_obj.qc_exist()
            f.write("{}\n".format(".".join([shotcode,step])))
            for key in qc_obj.detail:
                f.write("\t{key} : {content}\n".format(key=key, content=qc_obj.detail[key]))

    # create deliveried shot deli:
    file_maya_deli = create_delivered_shot(lst_scene_deliveredShot, "maya")
    file_mov_deli = create_delivered_shot(lst_mov_deliveredShot, "movie", is_copy_maya)

    # filter last version
    last_version_in_maya_result(file_maya_deli)
    last_version_in_mov_result(file_mov_deli)


    # todo: back up all information

    xls_bk = clarity.Template("xls_backup", DELI_SHOT_XLS_BK)
    clone_keyword = CLIENT_KEYWORD.copy()
    clone_keyword.update(date=datetime.now().strftime("%Y%m%d_%H%M%S"))
    try:
        bk_file_xls = xls_bk.format(clone_keyword, expand_var=True)

        if not osp.isdir(osp.dirname(bk_file_xls)):
            os.makedirs(osp.dirname(bk_file_xls))
        dct_c = {}
        for root, dirs, files in os.walk(osp.dirname(osp.dirname(bk_file_xls))):
            for file in files:
                if file.endswith(".ma"):
                    eps, seq, shot, _, _, version = file.split("_")[1:-1]
                    shot_code = "_".join([eps, str(int(seq)).zfill(2), shot])
                    dct_c.update({shot_code:version})
        create_xls_file(bk_file_xls, dct_c)
        # end todo
    except Exception as e:
        raw_input("Don't have any scenes which is deliveried,({}), please press enter to continue".format(str(e)))
    # remove temp:
    if osp.isdir(FOLDER_TEMP):
        shutil.rmtree(FOLDER_TEMP)

    os.system(log_file)

    raw_input("Please press enter to exit.")


if __name__ == "__main__":
    run()
